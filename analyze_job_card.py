"""
Deep analysis of Job Card Template
Focus on evaluation grid, cutter map, and business logic
"""

import openpyxl
from pathlib import Path

def analyze_job_card():
    file_path = Path(r"D:\PycharmProjects\Work Files\2025-ARDT-LV4-015-14204328.xlsx")

    print("="*80)
    print("JOB CARD TEMPLATE - DETAILED ANALYSIS")
    print("="*80)
    print()

    wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)

    # Focus on key sheets for job card logic
    key_sheets = [
        'Data',  # Main data entry
        'Evaluation',  # Cutter evaluation grid
        'E checklist',  # Evaluation checklist
        'ARDT Cutter Entry',  # Cutter data entry
        'Eng. Cutter Entry',  # Engineering cutter entry
        'Eval-LSTK',  # LSTK customer evaluation
        'Eval & Quot-AR',  # ARAMCO evaluation and quote
        'Instructions',  # Technical instructions
        'Router Sheet',  # Process routing
        'API Thread Inspection',  # API thread inspection
        'Die Check Entry',  # Die check / NDT
        'Quotation',  # Quotation sheet
    ]

    for sheet_name in key_sheets:
        if sheet_name in wb.sheetnames:
            print(f"\n{'#'*80}")
            print(f"# SHEET: {sheet_name}")
            print(f"{'#'*80}\n")
            analyze_detailed_sheet(wb, sheet_name)

    wb.close()

def analyze_detailed_sheet(wb, sheet_name):
    ws = wb[sheet_name]

    print(f"**Purpose:** {guess_sheet_purpose(sheet_name)}\n")
    print(f"**Dimensions:** {ws.max_row} rows x {ws.max_column} columns\n")

    # Special handling for evaluation grids
    if 'eval' in sheet_name.lower():
        analyze_evaluation_grid(ws, sheet_name)
    elif 'cutter entry' in sheet_name.lower():
        analyze_cutter_entry(ws, sheet_name)
    elif sheet_name == 'Data':
        analyze_data_sheet(ws)
    else:
        analyze_general_sheet(ws)


def guess_sheet_purpose(sheet_name):
    purposes = {
        'Data': 'Master data entry sheet - contains bit identification, MAT numbers, and references to other sheets',
        'Evaluation': 'Cutter-by-cutter evaluation grid with symbols (X, O, S, R, L, V, P, I, B)',
        'E checklist': 'Evaluation checklist for QC procedures',
        'ARDT Cutter Entry': 'Data entry for ARDT cutter replacements',
        'Eng. Cutter Entry': 'Engineering-approved cutter replacements',
        'Eval-LSTK': 'LSTK customer-specific evaluation form',
        'Eval & Quot-AR': 'ARAMCO evaluation and quotation',
        'Instructions': 'Technical repair/evaluation instructions',
        'Router Sheet': 'Process routing and workflow tracking',
        'API Thread Inspection': 'API thread inspection checklist',
        'Die Check Entry': 'Die check and NDT inspection data',
        'Quotation': 'Cost quotation and pricing',
    }
    return purposes.get(sheet_name, 'Unknown purpose')


def analyze_evaluation_grid(ws, sheet_name):
    """Analyze evaluation grid structure - cutter map layout"""
    print("**EVALUATION GRID ANALYSIS:**\n")

    # Look for evaluation symbols and grid structure
    symbols_found = set()
    blade_structure = []

    # Sample central area where evaluation grids typically are
    print("**Sample Grid Area (Rows 5-25, Cols A-T):**\n")

    for row in range(5, min(26, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(21, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            value = cell.value

            # Check for evaluation symbols
            if value in ['X', 'O', 'S', 'R', 'L', 'V', 'P', 'I', 'B', 'x', 'o', 's', 'r', 'l', 'v', 'p', 'i', 'b']:
                symbols_found.add(value.upper())

            # Truncate for display
            if isinstance(value, str) and len(value) > 15:
                value = value[:12] + "..."

            row_data.append(str(value) if value is not None else "")

        print(f"  Row {row}: {' | '.join(row_data[:10])}")

    print()
    if symbols_found:
        print(f"**Evaluation Symbols Found:** {', '.join(sorted(symbols_found))}\n")
    else:
        print("**Evaluation Symbols:** None detected in sampled area (may be in different location)\n")

    # Look for blade/pocket labels
    print("**Blade/Pocket Structure (sampling):**\n")
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, min(31, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            value = str(cell.value).lower() if cell.value else ""

            if 'blade' in value or 'pocket' in value or 'cutter' in value and 'position' in value:
                print(f"  Row {row}, Col {col}: '{cell.value}'")
                if len(blade_structure) < 10:  # Limit output
                    blade_structure.append((row, col, cell.value))

    if not blade_structure:
        print("  (No explicit blade/pocket labels found in sampled area)")

    print()


def analyze_cutter_entry(ws, sheet_name):
    """Analyze cutter entry sheets"""
    print("**CUTTER ENTRY ANALYSIS:**\n")

    # Find header row
    header_row = None
    for row in range(1, min(21, ws.max_row + 1)):
        for col in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            value = str(cell.value).lower() if cell.value else ""
            if any(keyword in value for keyword in ['cutter', 'qty', 'quantity', 'part #', 'blade', 'pocket']):
                header_row = row
                break
        if header_row:
            break

    if header_row:
        print(f"**Detected Header Row:** Row {header_row}\n")
        print("**Headers:**")
        for col in range(1, min(21, ws.max_column + 1)):
            cell = ws.cell(row=header_row, column=col)
            if cell.value:
                print(f"  Col {col}: {cell.value}")
        print()

        # Sample data rows
        print(f"**Sample Data Rows ({header_row+1} to {header_row+10}):**\n")
        for row in range(header_row + 1, min(header_row + 11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if isinstance(value, str) and len(value) > 20:
                    value = value[:17] + "..."
                row_data.append(str(value) if value is not None else "")
            print(f"  Row {row}: {' | '.join(row_data)}")
    else:
        print("**Header row not automatically detected.** Showing first 15 rows:\n")
        for row in range(1, min(16, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(8, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if isinstance(value, str) and len(value) > 20:
                    value = value[:17] + "..."
                row_data.append(str(value) if value is not None else "")
            print(f"  Row {row}: {' | '.join(row_data)}")

    print()


def analyze_data_sheet(ws):
    """Analyze main Data sheet - the control center"""
    print("**DATA SHEET CONTROL CENTER:**\n")

    # Key fields in Data sheet (typically in column A-B)
    print("**Key Bit Identification Fields (Cols A-B, Rows 1-20):**\n")
    for row in range(1, min(21, ws.max_row + 1)):
        col_a = ws.cell(row=row, column=1).value  # Label
        col_b = ws.cell(row=row, column=2).value  # Value
        col_b_formula = ws.cell(row=row, column=2).value if ws.cell(row=row, column=2).data_type == 'f' else None

        if col_a:
            if col_b_formula:
                print(f"  Row {row}: {col_a} = FORMULA: {col_b_formula}")
            else:
                print(f"  Row {row}: {col_a} = {col_b}")

    print()

    # Look for cross-sheet formulas
    print("**Cross-Sheet Formula References (sampling first 50 rows):**\n")
    cross_sheet_refs = []
    for row in range(1, min(51, ws.max_row + 1)):
        for col in range(1, min(21, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f' and cell.value:
                formula = str(cell.value)
                # Check for sheet references
                if "!" in formula:
                    sheets_referenced = []
                    for other_sheet in ['Evaluation', 'ARDT Cutter Entry', 'Eng. Cutter Entry',
                                       'Eval-LSTK', 'Instructions', 'Quotation', 'Rework']:
                        if other_sheet in formula:
                            sheets_referenced.append(other_sheet)

                    if sheets_referenced and len(cross_sheet_refs) < 15:
                        cross_sheet_refs.append({
                            'row': row,
                            'col': col,
                            'formula_start': formula[:80],
                            'sheets': sheets_referenced
                        })

    for ref in cross_sheet_refs:
        print(f"  Row {ref['row']}, Col {ref['col']}: {ref['formula_start']}...")
        print(f"    → References: {', '.join(ref['sheets'])}")

    print()


def analyze_general_sheet(ws):
    """General analysis for other sheets"""
    print("**GENERAL STRUCTURE:**\n")

    # First 20 rows, first 10 columns
    print("**First 20 rows x 10 columns:**\n")
    for row in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(11, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            value = cell.value

            if cell.data_type == 'f':
                value = f"={{formula}}"
            elif isinstance(value, str) and len(value) > 18:
                value = value[:15] + "..."

            row_data.append(str(value) if value is not None else "")

        print(f"  Row {row}: {' | '.join(row_data)}")

    print()


if __name__ == "__main__":
    analyze_job_card()
