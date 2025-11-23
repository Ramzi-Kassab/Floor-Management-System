"""
Data export utilities for PDF and Excel formats

Provides:
- PDF report generation
- Excel spreadsheet export
- CSV export utilities
- Formatted reports
"""
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone


# ============================================================================
# EXCEL EXPORT UTILITIES
# ============================================================================

class ExcelExporter:
    """
    Utility class for exporting data to Excel format

    Usage:
        exporter = ExcelExporter('Employee Report')
        exporter.add_sheet('Employees', headers, data)
        return exporter.get_response('employees.xlsx')
    """

    def __init__(self, title='Report'):
        """Initialize the Excel exporter"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            self.openpyxl = openpyxl
            self.Font = Font
            self.Alignment = Alignment
            self.PatternFill = PatternFill
            self.Border = Border
            self.Side = Side
        except ImportError:
            raise ImportError('openpyxl is required for Excel export. Install it with: pip install openpyxl')

        self.workbook = self.openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.title = title

    def add_sheet(self, sheet_name, headers, data, column_widths=None):
        """
        Add a sheet to the workbook

        Args:
            sheet_name: Name of the sheet
            headers: List of column headers
            data: List of rows (each row is a list of values)
            column_widths: Optional dict of column widths {column_letter: width}
        """
        worksheet = self.workbook.create_sheet(title=sheet_name)

        # Header styling
        header_fill = self.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = self.Font(color='FFFFFF', bold=True, size=12)
        header_alignment = self.Alignment(horizontal='center', vertical='center')

        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = self.Alignment(vertical='top', wrap_text=True)

        # Set column widths
        if column_widths:
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
        else:
            # Auto-width based on headers
            for col_idx, header in enumerate(headers, start=1):
                col_letter = self.openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = len(str(header)) + 5

        # Freeze header row
        worksheet.freeze_panes = 'A2'

        # Add borders
        thin_border = self.Border(
            left=self.Side(style='thin'),
            right=self.Side(style='thin'),
            top=self.Side(style='thin'),
            bottom=self.Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=len(data) + 1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

    def add_summary_sheet(self, sheet_name, summary_data):
        """
        Add a summary sheet with key-value pairs

        Args:
            sheet_name: Name of the sheet
            summary_data: Dict of {label: value} pairs
        """
        worksheet = self.workbook.create_sheet(title=sheet_name)

        # Styling
        label_font = self.Font(bold=True, size=11)
        label_fill = self.PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        row = 1
        for label, value in summary_data.items():
            # Label cell
            label_cell = worksheet.cell(row=row, column=1, value=label)
            label_cell.font = label_font
            label_cell.fill = label_fill

            # Value cell
            value_cell = worksheet.cell(row=row, column=2, value=value)

            row += 1

        # Set column widths
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 40

    def get_response(self, filename):
        """
        Get HTTP response with the Excel file

        Args:
            filename: Name of the file to download

        Returns:
            HttpResponse with Excel file
        """
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def save_to_file(self, filepath):
        """Save the workbook to a file"""
        self.workbook.save(filepath)


# ============================================================================
# PDF EXPORT UTILITIES
# ============================================================================

class PDFExporter:
    """
    Utility class for exporting data to PDF format

    Usage:
        exporter = PDFExporter('Employee Report')
        exporter.add_title('Employee Directory')
        exporter.add_table(headers, data)
        return exporter.get_response('employees.pdf')
    """

    def __init__(self, title='Report'):
        """Initialize the PDF exporter"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

            self.SimpleDocTemplate = SimpleDocTemplate
            self.Table = Table
            self.TableStyle = TableStyle
            self.Paragraph = Paragraph
            self.Spacer = Spacer
            self.PageBreak = PageBreak
            self.colors = colors
            self.inch = inch
            self.letter = letter
            self.A4 = A4
            self.TA_CENTER = TA_CENTER
            self.TA_LEFT = TA_LEFT
            self.TA_RIGHT = TA_RIGHT
        except ImportError:
            raise ImportError('reportlab is required for PDF export. Install it with: pip install reportlab')

        self.title = title
        self.elements = []
        self.styles = self._get_styles()
        self.buffer = BytesIO()

    def _get_styles(self):
        """Get paragraph styles"""
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        styles = getSampleStyleSheet()

        # Custom styles
        styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=self.colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=self.TA_CENTER
        ))

        styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=self.colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        ))

        return styles

    def add_title(self, title_text):
        """Add a title to the document"""
        title = self.Paragraph(title_text, self.styles['CustomTitle'])
        self.elements.append(title)
        self.elements.append(self.Spacer(1, 0.2 * self.inch))

    def add_heading(self, heading_text):
        """Add a heading to the document"""
        heading = self.Paragraph(heading_text, self.styles['CustomHeading'])
        self.elements.append(heading)

    def add_paragraph(self, text, style_name='Normal'):
        """Add a paragraph to the document"""
        para = self.Paragraph(text, self.styles[style_name])
        self.elements.append(para)
        self.elements.append(self.Spacer(1, 0.1 * self.inch))

    def add_table(self, headers, data, col_widths=None):
        """
        Add a table to the document

        Args:
            headers: List of column headers
            data: List of rows (each row is a list of values)
            col_widths: Optional list of column widths
        """
        # Prepare table data
        table_data = [headers] + data

        # Create table
        if col_widths:
            table = self.Table(table_data, colWidths=col_widths)
        else:
            table = self.Table(table_data)

        # Style the table
        style = self.TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), self.colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), self.colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), self.colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, self.colors.grey),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [self.colors.white, self.colors.HexColor('#F0F0F0')]),
        ])

        table.setStyle(style)
        self.elements.append(table)
        self.elements.append(self.Spacer(1, 0.2 * self.inch))

    def add_key_value_table(self, data_dict):
        """
        Add a key-value table to the document

        Args:
            data_dict: Dictionary of {key: value} pairs
        """
        table_data = [[key, str(value)] for key, value in data_dict.items()]

        table = self.Table(table_data, colWidths=[3 * self.inch, 4 * self.inch])

        style = self.TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), self.colors.HexColor('#E7E6E6')),
            ('TEXTCOLOR', (0, 0), (0, -1), self.colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, self.colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ])

        table.setStyle(style)
        self.elements.append(table)
        self.elements.append(self.Spacer(1, 0.2 * self.inch))

    def add_page_break(self):
        """Add a page break"""
        self.elements.append(self.PageBreak())

    def add_spacer(self, height=0.2):
        """Add vertical space"""
        self.elements.append(self.Spacer(1, height * self.inch))

    def build(self, page_size='letter'):
        """Build the PDF document"""
        size = self.letter if page_size == 'letter' else self.A4

        doc = self.SimpleDocTemplate(
            self.buffer,
            pagesize=size,
            rightMargin=0.75 * self.inch,
            leftMargin=0.75 * self.inch,
            topMargin=0.75 * self.inch,
            bottomMargin=0.75 * self.inch,
            title=self.title
        )

        # Add header and footer
        def add_page_number(canvas, doc):
            """Add page number to each page"""
            page_num = canvas.getPageNumber()
            text = f"Page {page_num}"
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.drawRightString(8 * self.inch, 0.5 * self.inch, text)
            canvas.drawString(0.75 * self.inch, 0.5 * self.inch, self.title)
            canvas.drawString(0.75 * self.inch, size[1] - 0.5 * self.inch,
                            f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
            canvas.restoreState()

        doc.build(self.elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

    def get_response(self, filename):
        """
        Get HTTP response with the PDF file

        Args:
            filename: Name of the file to download

        Returns:
            HttpResponse with PDF file
        """
        self.build()
        self.buffer.seek(0)

        response = HttpResponse(self.buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def save_to_file(self, filepath):
        """Save the PDF to a file"""
        self.build()
        with open(filepath, 'wb') as f:
            f.write(self.buffer.getvalue())


# ============================================================================
# CSV EXPORT UTILITIES
# ============================================================================

class CSVExporter:
    """
    Utility class for exporting data to CSV format

    Usage:
        exporter = CSVExporter()
        exporter.add_row(['Name', 'Email', 'Department'])
        exporter.add_rows(data)
        return exporter.get_response('employees.csv')
    """

    def __init__(self):
        """Initialize the CSV exporter"""
        import csv
        self.csv = csv
        self.buffer = BytesIO()
        self.writer = csv.writer(self.buffer.io if hasattr(self.buffer, 'io') else self.buffer)
        self.rows = []

    def add_row(self, row_data):
        """Add a single row"""
        self.rows.append(row_data)

    def add_rows(self, rows_data):
        """Add multiple rows"""
        self.rows.extend(rows_data)

    def get_response(self, filename):
        """
        Get HTTP response with the CSV file

        Args:
            filename: Name of the file to download

        Returns:
            HttpResponse with CSV file
        """
        import csv

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        for row in self.rows:
            writer.writerow(row)

        return response

    def save_to_file(self, filepath):
        """Save the CSV to a file"""
        import csv

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in self.rows:
                writer.writerow(row)


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def export_queryset_to_excel(queryset, filename, sheet_name='Data', fields=None, headers=None):
    """
    Export a Django queryset to Excel

    Usage:
        return export_queryset_to_excel(
            Employee.objects.all(),
            'employees.xlsx',
            fields=['employee_code', 'person__first_name', 'department__name'],
            headers=['Code', 'Name', 'Department']
        )

    Args:
        queryset: Django queryset
        filename: Output filename
        sheet_name: Name of the Excel sheet
        fields: List of field names to export (use __ for related fields)
        headers: List of column headers (if None, uses field names)

    Returns:
        HttpResponse with Excel file
    """
    # Get field values
    if fields:
        values = queryset.values_list(*fields)
        if not headers:
            headers = fields
    else:
        # Export all fields
        model = queryset.model
        fields = [f.name for f in model._meta.fields]
        values = queryset.values_list(*fields)
        headers = fields

    # Convert to list
    data = list(values)

    # Create Excel file
    exporter = ExcelExporter(f'{sheet_name} Report')
    exporter.add_sheet(sheet_name, headers, data)

    return exporter.get_response(filename)


def export_queryset_to_pdf(queryset, filename, title, fields=None, headers=None):
    """
    Export a Django queryset to PDF

    Args:
        queryset: Django queryset
        filename: Output filename
        title: PDF title
        fields: List of field names to export
        headers: List of column headers

    Returns:
        HttpResponse with PDF file
    """
    # Get field values
    if fields:
        values = queryset.values_list(*fields)
        if not headers:
            headers = fields
    else:
        model = queryset.model
        fields = [f.name for f in model._meta.fields]
        values = queryset.values_list(*fields)
        headers = fields

    # Convert to list and limit for PDF
    data = [[str(val)[:50] for val in row] for row in values[:100]]  # Limit to 100 rows for PDF

    # Create PDF
    exporter = PDFExporter(title)
    exporter.add_title(title)
    exporter.add_paragraph(f'Generated on {timezone.now().strftime("%B %d, %Y at %I:%M %p")}')
    exporter.add_paragraph(f'Total records: {queryset.count()}')
    exporter.add_table(headers, data)

    return exporter.get_response(filename)


def export_queryset_to_csv(queryset, filename, fields=None, headers=None):
    """
    Export a Django queryset to CSV

    Args:
        queryset: Django queryset
        filename: Output filename
        fields: List of field names to export
        headers: List of column headers

    Returns:
        HttpResponse with CSV file
    """
    # Get field values
    if fields:
        values = queryset.values_list(*fields)
        if not headers:
            headers = fields
    else:
        model = queryset.model
        fields = [f.name for f in model._meta.fields]
        values = queryset.values_list(*fields)
        headers = fields

    # Create CSV
    exporter = CSVExporter()
    exporter.add_row(headers)
    exporter.add_rows(values)

    return exporter.get_response(filename)
