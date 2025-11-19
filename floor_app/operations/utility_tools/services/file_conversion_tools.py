"""
File Conversion Tools Service

Convert between different file formats (PDF, Word, Excel, CSV, etc.)
"""

import io
import csv
import json
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class FileConversionService:
    """
    Service for file format conversions.
    """

    @classmethod
    def excel_to_csv(
        cls,
        excel_file,
        sheet_name: str = None,
        delimiter: str = ','
    ) -> io.StringIO:
        """
        Convert Excel file to CSV.

        Args:
            excel_file: Excel file object
            sheet_name: Specific sheet to convert (default: active sheet)
            delimiter: CSV delimiter

        Returns:
            StringIO object containing CSV data

        Example:
            csv_data = FileConversionService.excel_to_csv(
                excel_file=uploaded_file,
                sheet_name='Sheet1'
            )
        """
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter)

        for row in sheet.iter_rows(values_only=True):
            # Convert None to empty string
            row_values = [str(cell) if cell is not None else '' for cell in row]
            writer.writerow(row_values)

        output.seek(0)
        return output

    @classmethod
    def csv_to_excel(
        cls,
        csv_file,
        sheet_name: str = 'Sheet1',
        has_header: bool = True,
        delimiter: str = ','
    ) -> io.BytesIO:
        """
        Convert CSV file to Excel.

        Args:
            csv_file: CSV file object
            sheet_name: Name for the Excel sheet
            has_header: First row is header
            delimiter: CSV delimiter

        Returns:
            BytesIO object containing Excel file

        Example:
            excel_data = FileConversionService.csv_to_excel(
                csv_file=uploaded_file,
                sheet_name='Data',
                has_header=True
            )
        """
        # Read CSV
        csv_content = csv_file.read()
        if isinstance(csv_content, bytes):
            csv_content = csv_content.decode('utf-8')

        csv_reader = csv.reader(io.StringIO(csv_content), delimiter=delimiter)

        # Create Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # Write data
        for row_idx, row in enumerate(csv_reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)

                # Style header row
                if has_header and row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @classmethod
    def json_to_excel(
        cls,
        json_data: str | dict | list,
        sheet_name: str = 'Data'
    ) -> io.BytesIO:
        """
        Convert JSON data to Excel.

        Args:
            json_data: JSON string or dict/list
            sheet_name: Name for the Excel sheet

        Returns:
            BytesIO object containing Excel file

        Example:
            excel_data = FileConversionService.json_to_excel(
                json_data='[{"name": "Ahmed", "age": 30}, {"name": "Sara", "age": 25}]',
                sheet_name='People'
            )
        """
        # Parse JSON if string
        if isinstance(json_data, str):
            data = json.loads(json_data)
        else:
            data = json_data

        # Create workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        if isinstance(data, list) and len(data) > 0:
            # List of dictionaries
            if isinstance(data[0], dict):
                # Write headers
                headers = list(data[0].keys())
                for col_idx, header in enumerate(headers, start=1):
                    cell = sheet.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

                # Write data
                for row_idx, item in enumerate(data, start=2):
                    for col_idx, header in enumerate(headers, start=1):
                        value = item.get(header, '')
                        sheet.cell(row=row_idx, column=col_idx, value=str(value))

            # List of lists
            else:
                for row_idx, row_data in enumerate(data, start=1):
                    if isinstance(row_data, (list, tuple)):
                        for col_idx, value in enumerate(row_data, start=1):
                            sheet.cell(row=row_idx, column=col_idx, value=str(value))
                    else:
                        sheet.cell(row=row_idx, column=1, value=str(row_data))

        elif isinstance(data, dict):
            # Write key-value pairs
            sheet.cell(row=1, column=1, value='Key').font = Font(bold=True)
            sheet.cell(row=1, column=2, value='Value').font = Font(bold=True)

            for row_idx, (key, value) in enumerate(data.items(), start=2):
                sheet.cell(row=row_idx, column=1, value=str(key))
                sheet.cell(row=row_idx, column=2, value=str(value))

        # Auto-size columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Save
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @classmethod
    def excel_to_json(
        cls,
        excel_file,
        sheet_name: str = None
    ) -> str:
        """
        Convert Excel file to JSON.

        Args:
            excel_file: Excel file object
            sheet_name: Specific sheet to convert

        Returns:
            JSON string

        Example:
            json_data = FileConversionService.excel_to_json(
                excel_file=uploaded_file,
                sheet_name='Sheet1'
            )
        """
        workbook = openpyxl.load_workbook(excel_file, data_only=True)

        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # Get all rows
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) == 0:
            return json.dumps([])

        # First row is headers
        headers = [str(h) if h is not None else f'Column{i}' for i, h in enumerate(rows[0])]

        # Convert to list of dicts
        data = []
        for row in rows[1:]:
            row_dict = {}
            for header, value in zip(headers, row):
                row_dict[header] = value if value is not None else ''
            data.append(row_dict)

        return json.dumps(data, indent=2, default=str)

    @classmethod
    def csv_to_json(
        cls,
        csv_file,
        has_header: bool = True,
        delimiter: str = ','
    ) -> str:
        """
        Convert CSV file to JSON.

        Args:
            csv_file: CSV file object
            has_header: First row is header
            delimiter: CSV delimiter

        Returns:
            JSON string

        Example:
            json_data = FileConversionService.csv_to_json(
                csv_file=uploaded_file,
                has_header=True
            )
        """
        csv_content = csv_file.read()
        if isinstance(csv_content, bytes):
            csv_content = csv_content.decode('utf-8')

        csv_reader = csv.reader(io.StringIO(csv_content), delimiter=delimiter)
        rows = list(csv_reader)

        if len(rows) == 0:
            return json.dumps([])

        if has_header:
            headers = rows[0]
            data = []
            for row in rows[1:]:
                row_dict = {}
                for header, value in zip(headers, row):
                    row_dict[header] = value
                data.append(row_dict)
        else:
            data = rows

        return json.dumps(data, indent=2)

    @classmethod
    def json_to_csv(
        cls,
        json_data: str | dict | list,
        delimiter: str = ','
    ) -> io.StringIO:
        """
        Convert JSON to CSV.

        Args:
            json_data: JSON string or dict/list
            delimiter: CSV delimiter

        Returns:
            StringIO object containing CSV data

        Example:
            csv_data = FileConversionService.json_to_csv(
                json_data='[{"name": "Ahmed", "age": 30}]'
            )
        """
        # Parse JSON if string
        if isinstance(json_data, str):
            data = json.loads(json_data)
        else:
            data = json_data

        output = io.StringIO()
        writer = csv.writer(output, delimiter=delimiter)

        if isinstance(data, list) and len(data) > 0:
            if isinstance(data[0], dict):
                # Write headers
                headers = list(data[0].keys())
                writer.writerow(headers)

                # Write data
                for item in data:
                    row = [item.get(header, '') for header in headers]
                    writer.writerow(row)
            else:
                # List of lists
                for row in data:
                    if isinstance(row, (list, tuple)):
                        writer.writerow(row)
                    else:
                        writer.writerow([row])

        elif isinstance(data, dict):
            # Write key-value pairs
            writer.writerow(['Key', 'Value'])
            for key, value in data.items():
                writer.writerow([key, value])

        output.seek(0)
        return output
