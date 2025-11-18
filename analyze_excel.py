"""
Excel Workbook Analysis Script
Analyzes Excel workbooks to understand structure, formulas, and business logic
WITHOUT reading all data (token-efficient analysis)
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os
from pathlib import Path

def analyze_workbook(file_path):
    """Analyze a single Excel workbook"""
    print(f"\n{'='*80}")
    print(f"## Workbook: {Path(file_path).name}")
    print(f"Path: {file_path}")
    print(f"{'='*80}\n")

    try:
        # Open in read-only mode for efficiency
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=False)

        print(f"**Number of sheets:** {len(wb.sheetnames)}\n")
        print(f"**Sheet names:** {', '.join(wb.sheetnames)}\n")

        # Analyze each sheet
        for sheet_name in wb.sheetnames:
            analyze_sheet(wb, sheet_name)

        wb.close()

    except PermissionError:
        print(f"WARNING: **File is currently OPEN in Excel** - Cannot analyze while open.\n")
        print(f"   Please close the file and re-run analysis.\n")
    except Exception as e:
        print(f"ERROR: **Error analyzing workbook:** {str(e)}\n")


def analyze_sheet(wb, sheet_name):
    """Analyze a single sheet - structure focused, not data-heavy"""
    print(f"\n### Sheet: `{sheet_name}`")
    print(f"{'-'*60}\n")

    ws = wb[sheet_name]

    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column

    print(f"**Dimensions:** {max_row} rows × {max_col} columns\n")

    # Read header row(s) - typically first 1-2 rows
    print(f"**Header Row (Row 1):**")
    headers = []
    for col in range(1, min(max_col + 1, 51)):  # Limit to 50 columns
        cell = ws.cell(row=1, column=col)
        if cell.value:
            headers.append(f"  - Col {get_column_letter(col)}: `{cell.value}`")

    if headers:
        print("\n".join(headers[:30]))  # Show first 30 headers
        if len(headers) > 30:
            print(f"  ... and {len(headers) - 30} more columns")
    else:
        print("  (No headers found in row 1)")

    print()

    # Sample first 10-15 data rows (starting from row 2)
    print(f"**Sample Data (Rows 2-12, first 10 columns):**\n")
    sample_data = []
    for row in range(2, min(13, max_row + 1)):  # Rows 2-12
        row_data = []
        for col in range(1, min(11, max_col + 1)):  # First 10 columns
            cell = ws.cell(row=row, column=col)
            value = cell.value

            # Truncate long values
            if isinstance(value, str) and len(value) > 40:
                value = value[:37] + "..."

            row_data.append(str(value) if value is not None else "")

        sample_data.append(f"  Row {row}: {' | '.join(row_data)}")

    if sample_data:
        print("\n".join(sample_data[:10]))
    else:
        print("  (No data rows found)")

    print()

    # Analyze formulas in header row and first few data rows
    print(f"**Formula Analysis (sampling rows 1-20):**\n")
    formula_columns = []

    for col in range(1, min(max_col + 1, 51)):
        col_letter = get_column_letter(col)
        formulas_found = []

        # Check first 20 rows for formulas in this column
        for row in range(1, min(21, max_row + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.data_type == 'f':  # Formula cell
                formula = cell.value
                if formula and formula not in formulas_found:
                    formulas_found.append(formula)
                    if len(formulas_found) <= 2:  # Show first 2 unique formulas
                        formula_columns.append(
                            f"  - Col {col_letter} (Row {row}): `{formula}`"
                        )

    if formula_columns:
        print("\n".join(formula_columns[:20]))  # Show first 20 formula examples
        if len(formula_columns) > 20:
            print(f"  ... and {len(formula_columns) - 20} more formula examples")
    else:
        print("  (No formulas detected in sampled range)")

    print()

    # Check for named ranges
    print(f"**Named Ranges in Workbook:**\n")
    try:
        if hasattr(wb, 'defined_names') and wb.defined_names:
            named_list = list(wb.defined_names)[:20]  # First 20
            for name in named_list:
                dn = wb.defined_names[name]
                print(f"  - `{name}`: {dn.value if hasattr(dn, 'value') else dn}")
            if len(wb.defined_names) > 20:
                print(f"  ... and {len(wb.defined_names) - 20} more")
        else:
            print("  (No named ranges found)")
    except Exception as e:
        print(f"  (Error reading named ranges: {e})")

    print()

    # Check for data validation
    print(f"**Data Validation (sampling first 100 cells):**\n")
    validation_found = False
    for col in range(1, min(11, max_col + 1)):
        for row in range(1, min(11, max_row + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.data_validation and cell.data_validation.type:
                validation_found = True
                print(f"  - {get_column_letter(col)}{row}: {cell.data_validation.type}")

    if not validation_found:
        print("  (No data validation detected in sampled range)")

    print("\n" + "="*60 + "\n")


def main():
    """Main analysis function"""
    work_files_dir = Path(r"D:\PycharmProjects\Work Files")

    # Find all Excel files (excluding temp files)
    excel_files = [
        f for f in work_files_dir.glob("*.xlsx")
        if not f.name.startswith("~$")
    ]

    print(f"\n{'#'*80}")
    print(f"# EXCEL WORKBOOK ANALYSIS REPORT")
    print(f"# Directory: {work_files_dir}")
    print(f"# Files found: {len(excel_files)}")
    print(f"{'#'*80}\n")

    # Sort by name for consistent ordering
    excel_files.sort(key=lambda x: x.name)

    for excel_file in excel_files:
        analyze_workbook(str(excel_file))

    print(f"\n{'#'*80}")
    print(f"# ANALYSIS COMPLETE")
    print(f"{'#'*80}\n")


if __name__ == "__main__":
    main()
